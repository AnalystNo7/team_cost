"""
Export Service for Excel and PDF reports
"""
import io
import os
from decimal import Decimal
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from app.schemas.cost import CostCalculationResult

# Register Cyrillic font
_font_registered = False

def register_cyrillic_font():
    """Register a font that supports Cyrillic characters"""
    global _font_registered
    if _font_registered:
        return

    # Try to find DejaVuSans font (common on Linux)
    font_paths = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/dejavu/DejaVuSans.ttf',
        'C:/Windows/Fonts/arial.ttf',
        'C:/Windows/Fonts/calibri.ttf',
        '/System/Library/Fonts/Helvetica.ttc',
    ]

    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                _font_registered = True
                return
            except Exception:
                continue

    # If no font found, we'll use default (may not display Cyrillic correctly)
    _font_registered = True


class ExportService:
    @staticmethod
    def export_to_excel(result: CostCalculationResult) -> bytes:
        """Export calculation result to Excel format"""
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Сводка"
        ExportService._create_summary_sheet(ws_summary, result)

        # Stages detail sheet
        ws_stages = wb.create_sheet("По этапам")
        ExportService._create_stages_sheet(ws_stages, result)

        # Monthly breakdown sheet
        ws_monthly = wb.create_sheet("По месяцам")
        ExportService._create_monthly_sheet(ws_monthly, result)

        # Roles breakdown sheet
        ws_roles = wb.create_sheet("По ролям")
        ExportService._create_roles_sheet(ws_roles, result)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _create_summary_sheet(ws: Any, result: CostCalculationResult):
        """Create summary sheet"""
        header_font = Font(bold=True, size=14)
        label_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws['A1'] = f"Расчёт стоимости проекта: {result.calculation_name}"
        ws['A1'].font = header_font
        ws.merge_cells('A1:D1')

        rows = [
            ("", ""),
            ("Методология:", result.methodology.upper()),
            ("Период:", f"{result.start_date} — {result.end_date}"),
            ("", ""),
            ("ИТОГОВЫЕ ПОКАЗАТЕЛИ", ""),
            ("Общая себестоимость:", f"{result.total_cost:,.2f} ₽"),
            ("  - Внутренние ресурсы:", f"{result.total_internal_cost:,.2f} ₽"),
            ("  - Внешние ресурсы:", f"{result.total_external_cost:,.2f} ₽"),
            ("Общая выручка:", f"{result.total_revenue:,.2f} ₽"),
            ("Маржа:", f"{result.total_margin:,.2f} ₽"),
            ("Маржинальность:", f"{result.margin_percent}%" if result.margin_percent else "N/A"),
        ]

        for i, (label, value) in enumerate(rows, start=2):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            if label and not label.startswith(" "):
                ws[f'A{i}'].font = label_font

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

    @staticmethod
    def _create_stages_sheet(ws: Any, result: CostCalculationResult):
        """Create stages breakdown sheet"""
        headers = ["Этап", "Период", "Себестоимость", "Выручка", "Маржа", "Маржа %"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, stage in enumerate(result.stages, 2):
            ws.cell(row=row, column=1, value=stage.stage_name)
            ws.cell(row=row, column=2, value=f"{stage.start_month} — {stage.end_month}")
            ws.cell(row=row, column=3, value=float(stage.total_cost))
            ws.cell(row=row, column=4, value=float(stage.total_revenue))
            ws.cell(row=row, column=5, value=float(stage.total_margin))
            ws.cell(row=row, column=6, value=float(stage.margin_percent) if stage.margin_percent else 0)

        # Format columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    @staticmethod
    def _create_monthly_sheet(ws: Any, result: CostCalculationResult):
        """Create monthly breakdown sheet"""
        headers = ["Месяц", "Себестоимость", "Выручка", "Маржа"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        months = sorted(result.cost_by_month.keys())
        for row, month in enumerate(months, 2):
            cost = result.cost_by_month.get(month, Decimal("0"))
            revenue = result.revenue_by_month.get(month, Decimal("0"))
            margin = revenue - cost

            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=float(cost))
            ws.cell(row=row, column=3, value=float(revenue))
            ws.cell(row=row, column=4, value=float(margin))

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 18

    @staticmethod
    def _create_roles_sheet(ws: Any, result: CostCalculationResult):
        """Create roles breakdown sheet"""
        headers = ["Роль", "Себестоимость"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row, (role, cost) in enumerate(sorted(result.cost_by_role.items()), 2):
            ws.cell(row=row, column=1, value=role)
            ws.cell(row=row, column=2, value=float(cost))

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    @staticmethod
    def export_to_pdf(result: CostCalculationResult) -> bytes:
        """Export calculation result to PDF format"""
        register_cyrillic_font()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )

        styles = getSampleStyleSheet()

        # Try to use Cyrillic font if registered
        font_name = 'CyrillicFont' if _font_registered else 'Helvetica'

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=16,
            spaceAfter=20
        )

        heading2_style = ParagraphStyle(
            'CustomHeading2',
            parent=styles['Heading2'],
            fontName=font_name,
            fontSize=12,
            spaceAfter=10
        )

        elements = []

        # Title
        elements.append(Paragraph(f"Calculation: {result.calculation_name}", title_style))
        elements.append(Spacer(1, 10))

        # Summary table
        summary_data = [
            ["Parameter", "Value"],
            ["Methodology", result.methodology.upper()],
            ["Period", f"{result.start_date} - {result.end_date}"],
            ["Total Cost", f"{float(result.total_cost):,.2f} RUB"],
            ["Internal Resources", f"{float(result.total_internal_cost):,.2f} RUB"],
            ["External Resources", f"{float(result.total_external_cost):,.2f} RUB"],
            ["Total Revenue", f"{float(result.total_revenue):,.2f} RUB"],
            ["Margin", f"{float(result.total_margin):,.2f} RUB"],
            ["Margin %", f"{result.margin_percent}%" if result.margin_percent else "N/A"],
        ]

        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Stages table
        elements.append(Paragraph("Stages Breakdown", heading2_style))
        stages_data = [["Stage", "Period", "Cost", "Revenue", "Margin", "Margin %"]]
        for stage in result.stages:
            stages_data.append([
                stage.stage_name,
                f"{stage.start_month} - {stage.end_month}",
                f"{float(stage.total_cost):,.2f}",
                f"{float(stage.total_revenue):,.2f}",
                f"{float(stage.total_margin):,.2f}",
                f"{stage.margin_percent}%" if stage.margin_percent else "N/A"
            ])

        stages_table = Table(stages_data, colWidths=[150, 100, 100, 100, 100, 80])
        stages_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(stages_table)

        # Roles breakdown
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Roles Breakdown", heading2_style))
        roles_data = [["Role", "Cost"]]
        for role, cost in sorted(result.cost_by_role.items()):
            roles_data.append([role, f"{float(cost):,.2f}"])

        roles_table = Table(roles_data, colWidths=[250, 150])
        roles_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(roles_table)

        # Monthly breakdown
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Monthly Breakdown", heading2_style))
        monthly_data = [["Month", "Cost", "Revenue", "Margin"]]
        for month in sorted(result.cost_by_month.keys()):
            cost = float(result.cost_by_month.get(month, 0))
            revenue = float(result.revenue_by_month.get(month, 0))
            margin = revenue - cost
            monthly_data.append([
                month,
                f"{cost:,.2f}",
                f"{revenue:,.2f}",
                f"{margin:,.2f}"
            ])

        monthly_table = Table(monthly_data, colWidths=[100, 120, 120, 120])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(monthly_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
